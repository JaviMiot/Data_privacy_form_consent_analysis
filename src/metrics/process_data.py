from enum import StrEnum
import spacy
import pandas as pd
import openpyxl
import openpyxl.utils

from src.metrics.mer_trans_metrics.calculator import MerTrans
from src.metrics.mer_trans_metrics.models import MerTransMetrics
from src.metrics.clears_metrics.calculator import ClearsCalculator
from src.metrics.readability_complexity.calculator import index_calculator


class InputColumn(StrEnum):
    DOCUMENT_ID = "document_id"
    ORIGINAL_SENTENCE_ID = "original_sentence_id"
    ORIGINAL_TEXT = "original_text"
    SIMPLIFIED_TEXT = "simplified_text"
    PREDICTION_TEXT = "prediction_text"


class SummaryColumn(StrEnum):
    DOCUMENT_ID = "document_id"
    ORIG_BLEU = "orig_bleu"
    ORIG_SARI = "orig_sari"
    ORIG_BERTSCORE_PRECISION = "orig_bertscore_precision"
    ORIG_BERTSCORE_RECALL = "orig_bertscore_recall"
    ORIG_BERTSCORE_F1 = "orig_bertscore_f1"
    ORIG_MEANING_BERT = "orig_meaning_bert"
    ORIG_COS_TFIDF = "orig_cos_tfidf"
    ORIG_COS_EMBED = "orig_cos_embed"
    ORIG_COS_AVG = "orig_cos_avg"
    ORIG_ROBERTA_SENSE_FACIL_DOES_NOT_PRESERVE = "orig_roberta_sense_facil_does_not_preserve"
    ORIG_ROBERTA_SENSE_FACIL_PRESERVES_MEANING = "orig_roberta_sense_facil_preserves_meaning"
    GOLD_BLEU = "gold_bleu"
    GOLD_SARI = "gold_sari"
    GOLD_BERTSCORE_PRECISION = "gold_bertscore_precision"
    GOLD_BERTSCORE_RECALL = "gold_bertscore_recall"
    GOLD_BERTSCORE_F1 = "gold_bertscore_f1"
    GOLD_MEANING_BERT = "gold_meaning_bert"
    GOLD_ROBERTA_SENSE_FACIL_DOES_NOT_PRESERVE = "gold_roberta_sense_facil_does_not_preserve"
    GOLD_ROBERTA_SENSE_FACIL_PRESERVES_MEANING = "gold_roberta_sense_facil_preserves_meaning"



class MetricsProcessor:
    def __init__(self, data: pd.DataFrame, mer_trans: MerTrans, clears: ClearsCalculator):
        self.data = data
        self.mer_trans = mer_trans
        self.clears = clears
        self.nlp = spacy.load('es_core_news_sm')
        self.has_references = self._check_for_references()

    def _check_for_references(self) -> bool:
        """
        Determina si el dataset contiene referencias de expertos válidas.
        Se considera sin referencias si:
        1. La columna SIMPLIFIED_TEXT no existe.
        2. Toda la columna contiene solo '-' o está vacía.
        """
        if InputColumn.SIMPLIFIED_TEXT not in self.data.columns:
            return False
            
        # Verificar si hay contenido real más allá de '-' o NaN
        refs = self.data[InputColumn.SIMPLIFIED_TEXT].fillna("-").astype(str)
        valid_refs = refs[~refs.isin(["-", "", "nan", "None"])]
        
        return len(valid_refs) > 0

    def group_data_by_document_id(self):
        return self.data.sort_values(by=InputColumn.ORIGINAL_SENTENCE_ID).groupby(InputColumn.DOCUMENT_ID)

    def get_metrics_for_document(self, document: pd.DataFrame):
        sources = document[InputColumn.ORIGINAL_TEXT].tolist()
        predictions = document[InputColumn.PREDICTION_TEXT].tolist()
        
        if self.has_references:
            references = document[InputColumn.SIMPLIFIED_TEXT].tolist()
        else:
            # Fallback: Usar fuentes como referencias para medir preservación
            references = sources

        orig_metrics = self.mer_trans.get_all_metrics_origin(sources, references, predictions)
        gold_metrics = self.mer_trans.get_all_metrics_gold(sources, references, predictions)
        clears_metrics = self.clears.get_all_metrics(predictions, sources)

        return {
            "gold": gold_metrics,
            "orig": orig_metrics,
            "clears": clears_metrics
        }

    def calculate_mer_trans_metrics(self):
        data_by_document_id = self.group_data_by_document_id()
        data = data_by_document_id.apply(lambda x: self.get_metrics_for_document(x))
        return data

    def calculate_readability_metrics(self):
        data_by_document_id = self.group_data_by_document_id()
        
        def _get_readability(x):
            pred_text = " ".join(x[InputColumn.PREDICTION_TEXT].astype(str).tolist())
            orig_text = " ".join(x[InputColumn.ORIGINAL_TEXT].astype(str).tolist())
            
            pred = index_calculator(pred_text, self.nlp) if pred_text.strip() else {}
            orig = index_calculator(orig_text, self.nlp) if orig_text.strip() else {}
            
            simp = {}
            if self.has_references:
                simp_text = " ".join(x[InputColumn.SIMPLIFIED_TEXT].astype(str).tolist())
                simp = index_calculator(simp_text, self.nlp) if simp_text.strip() else {}
            
            return {
                "pred": pred if pred else {},
                "orig": orig if orig else {},
                "simp": simp if simp else {}
            }
            
        data = data_by_document_id.apply(_get_readability)
        return data

    def get_summary_table(self) -> pd.DataFrame:
        metrics_by_document = self.calculate_mer_trans_metrics()
        readability_by_document = self.calculate_readability_metrics()

        rows = []
        for document_id in metrics_by_document.index:
            metrics = metrics_by_document[document_id]
            origin_metrics = metrics["orig"]
            gold_metrics = metrics["gold"]
            readability = readability_by_document.get(document_id)

            row = {
                SummaryColumn.DOCUMENT_ID: document_id,
                SummaryColumn.ORIG_BLEU: round(origin_metrics.bleu.bleu, 4),
                SummaryColumn.ORIG_SARI: round(origin_metrics.sari.sari, 4),
                SummaryColumn.ORIG_BERTSCORE_PRECISION: round(origin_metrics.bertscore.mean_precision, 4),
                SummaryColumn.ORIG_BERTSCORE_RECALL: round(origin_metrics.bertscore.mean_recall, 4),
                SummaryColumn.ORIG_BERTSCORE_F1: round(origin_metrics.bertscore.mean_f1, 4),
                SummaryColumn.ORIG_MEANING_BERT: round(origin_metrics.meaning_bert.mean_score, 4),
                SummaryColumn.ORIG_COS_TFIDF: round(metrics["clears"].tfidf.mean_score, 4),
                SummaryColumn.ORIG_COS_EMBED: round(metrics["clears"].embedding.mean_score, 4),
                SummaryColumn.ORIG_COS_AVG: round(metrics["clears"].cosine_avg, 4),
                SummaryColumn.ORIG_ROBERTA_SENSE_FACIL_DOES_NOT_PRESERVE: round(origin_metrics.roberta_sense_facil.does_not_preserve, 4),
                SummaryColumn.ORIG_ROBERTA_SENSE_FACIL_PRESERVES_MEANING: round(origin_metrics.roberta_sense_facil.preserves_meaning, 4),
                SummaryColumn.GOLD_BLEU: round(gold_metrics.bleu.bleu, 4),
                SummaryColumn.GOLD_SARI: round(gold_metrics.sari.sari, 4),
                SummaryColumn.GOLD_BERTSCORE_PRECISION: round(gold_metrics.bertscore.mean_precision, 4),
                SummaryColumn.GOLD_BERTSCORE_RECALL: round(gold_metrics.bertscore.mean_recall, 4),
                SummaryColumn.GOLD_BERTSCORE_F1: round(gold_metrics.bertscore.mean_f1, 4),
                SummaryColumn.GOLD_MEANING_BERT: round(gold_metrics.meaning_bert.mean_score, 4),
                SummaryColumn.GOLD_ROBERTA_SENSE_FACIL_DOES_NOT_PRESERVE: round(gold_metrics.roberta_sense_facil.does_not_preserve, 4),
                SummaryColumn.GOLD_ROBERTA_SENSE_FACIL_PRESERVES_MEANING: round(gold_metrics.roberta_sense_facil.preserves_meaning, 4),
            }

            if readability:
                for prefix in ["pred", "orig", "simp"]:
                    if prefix == "simp" and not self.has_references:
                        continue
                        
                    read_metrics = readability.get(prefix, {})
                    key_prefix = "orig_" if prefix == "orig" else "simp_" if prefix == "simp" else ""
                    
                    for category, metrics_dict in read_metrics.items():
                        if isinstance(metrics_dict, dict):
                            for k, v in metrics_dict.items():
                                row[f"{key_prefix}{k}"] = v
                        else:
                            row[f"{key_prefix}{category}"] = read_metrics[category]

            rows.append(row)

        df = pd.DataFrame(rows)

        # Reorder columns: Original, Simplified, Prediction
        doc_id_col = SummaryColumn.DOCUMENT_ID.value
        ordered_cols = [doc_id_col]
        base_cols = []
        for col in df.columns:
            if col == doc_id_col:
                continue
            if col.startswith("orig_"):
                base = col[5:]
            elif col.startswith("simp_"):
                base = col[5:]
            else:
                base = col
                
            if base not in base_cols:
                base_cols.append(base)
                
        for base in base_cols:
            if f"orig_{base}" in df.columns:
                ordered_cols.append(f"orig_{base}")
            if f"simp_{base}" in df.columns and self.has_references:
                ordered_cols.append(f"simp_{base}")
            if base in df.columns:
                ordered_cols.append(base)
                
        df = df[ordered_cols]

        avg_data = {SummaryColumn.DOCUMENT_ID: "PROMEDIO"}
        for col in df.columns:
            if col != SummaryColumn.DOCUMENT_ID:
                if pd.api.types.is_numeric_dtype(df[col]):
                    avg_data[col] = round(df[col].mean(), 4)
                else:
                    avg_data[col] = ""

        avg_row = pd.DataFrame([avg_data])
        
        return pd.concat([df, avg_row], ignore_index=True)

    def save_consolidated_report(self, output_path: str):
        """
        Generates an Excel file with three sheets:
        1. Raw Data: The original CSV content.
        2. Summary by Document: Metrics per document + Average.
        3. Metrics Summary: Vertical averages grouped by category.
        """
        summary_df = self.get_summary_table()
        
        # Define categories
        lexical_keywords = ["bleu", "sari"]
        semantic_keywords = ["bertscore", "meaning_bert", "cos_", "roberta_sense_facil"]
        
        def get_category(col):
            if col == SummaryColumn.DOCUMENT_ID.value:
                return "Metadata"
            col_lower = col.lower()
            if any(k in col_lower for k in lexical_keywords):
                return "Lexical"
            if any(k in col_lower for k in semantic_keywords):
                return "Semantic"
            return "Readability"

        # Group columns for Sheet 2
        cols = list(summary_df.columns)
        metadata_cols = [c for c in cols if get_category(c) == "Metadata"]
        lexical_cols = [c for c in cols if get_category(c) == "Lexical"]
        semantic_cols = [c for c in cols if get_category(c) == "Semantic"]
        readability_cols = [c for c in cols if get_category(c) == "Readability"]
        
        ordered_cols = metadata_cols + lexical_cols + semantic_cols + readability_cols
        summary_df = summary_df[ordered_cols]

        # Prepare Sheet 3 (Vertical Summary)
        avg_row = summary_df[summary_df[SummaryColumn.DOCUMENT_ID.value] == "PROMEDIO"].iloc[0]
        vertical_data = []
        
        for category, cat_cols in [("Lexical", lexical_cols), ("Semantic", semantic_cols), ("Readability", readability_cols)]:
            cat_data = []
            for col in cat_cols:
                val = avg_row[col]
                # Skip if value is empty string, None or NaN
                if val == "" or pd.isna(val):
                    continue
                cat_data.append({
                    "Metric Label": col,
                    "Average Value": val
                })
            
            if cat_data:
                # Add category header
                vertical_data.append({"Metric Label": f"--- {category} ---", "Average Value": ""})
                vertical_data.extend(cat_data)
        
        metrics_summary_df = pd.DataFrame(vertical_data)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Raw Data
            self.data.to_excel(writer, sheet_name='Raw Data', index=False)
            
            # Sheet 2: Summary by Document
            summary_df.to_excel(writer, sheet_name='Summary by Document', index=False)
            
            # Sheet 3: Metrics Summary
            metrics_summary_df.to_excel(writer, sheet_name='Metrics Summary', index=False)
            
            # Formatting
            workbook = writer.book
            
            # Sheet 2 formatting
            sheet2 = writer.sheets['Summary by Document']
            for cell in sheet2[1]: # Header
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
            for i in range(len(summary_df.columns)):
                sheet2.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = 20

            # Sheet 1 formatting
            sheet1 = writer.sheets['Raw Data']
            for cell in sheet1[1]:
                cell.font = openpyxl.styles.Font(bold=True)

            # Sheet 3 formatting
            sheet3 = writer.sheets['Metrics Summary']
            # Header
            for cell in sheet3[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
            
            # Category headers and labels
            for row_idx, row in enumerate(vertical_data, start=2):
                label_cell = sheet3.cell(row=row_idx, column=2) # Column B in Excel (Metric Label)
                val_cell = sheet3.cell(row=row_idx, column=3)   # Column C in Excel (Average Value)
                
                label_text = str(row["Metric Label"])
                if label_text.startswith("---"):
                    label_cell.font = openpyxl.styles.Font(bold=True)
                    label_cell.fill = openpyxl.styles.PatternFill(start_color="E1E1E1", end_color="E1E1E1", fill_type="solid")
                else:
                    label_cell.font = openpyxl.styles.Font(bold=True)

            sheet3.column_dimensions['B'].width = 30
            sheet3.column_dimensions['C'].width = 20
            
        print(f"Reporte consolidado guardado en: {output_path}")

metrics_processor = lambda data, ngrams = 4: MetricsProcessor(data, MerTrans(ngrams), ClearsCalculator())
