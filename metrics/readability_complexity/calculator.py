from pylabeador import syllabify
import spacy
import silabeador
import requests
import pandas as pd #Necesita openpyxl -> pip install openpyxl
import openpyxl
import io
import textstat
import statistics
import re
import numpy
import os
from pathlib import Path



BASE_PATH = Path(__file__).parent
#En este fichero se calculan los indices de lectura facil, asi como su interpretacion

_MAPA_INDICES_CACHE = None
_ABREVIATURAS_CACHE = None
_PALABRAS_TODAS_CACHE = None
_FORMAS_5000_CACHE = None

def get_mapa_indices():
    global _MAPA_INDICES_CACHE
    if _MAPA_INDICES_CACHE is None:
        _MAPA_INDICES_CACHE = openpyxl.load_workbook(BASE_PATH / "Mapa_indices.xlsx")
    return _MAPA_INDICES_CACHE

def get_abreviaturas():
    global _ABREVIATURAS_CACHE
    if _ABREVIATURAS_CACHE is None:
        workbook = openpyxl.load_workbook(BASE_PATH / "lista_abreviaturas_esp.xlsx")
        worksheet = workbook.active
        _ABREVIATURAS_CACHE = []
        for cell in worksheet['A']:
            if cell.value is None: continue
            celda = str(cell.value).replace(".","\\.")
            celda = celda.replace('(', '\\(')
            celda = celda.replace(')', '\\)')
            pattern = r"(?<=[^a-zA-ZÀ-ú\d])(" + celda + r")(?=[^a-zA-ZÀ-ú\d])"
            replacement = str(worksheet['B'][cell.row-1].value) if worksheet['B'][cell.row-1].value else ""
            _ABREVIATURAS_CACHE.append((re.compile(pattern), replacement))
    return _ABREVIATURAS_CACHE

def get_palabras_todas():
    global _PALABRAS_TODAS_CACHE
    if _PALABRAS_TODAS_CACHE is None:
        ruta_archivo2 = os.path.join(os.path.dirname(__file__),"palabras_todas.txt")
        _PALABRAS_TODAS_CACHE = set()
        with open(ruta_archivo2, 'r', encoding='utf-8') as file2:
            for linea in file2:
                for palabra in linea.split():
                    _PALABRAS_TODAS_CACHE.add(palabra.lower())
    return _PALABRAS_TODAS_CACHE

def get_formas_5000_set():
    global _FORMAS_5000_CACHE
    if _FORMAS_5000_CACHE is None:
        url = 'https://corpus.rae.es/frec/5000_formas.txt'
        headers={
            'User-Agent': 'Mozilla/5.0',
            'From': 'youremail@domainexample'
        }
        try:
            r = requests.get(url, verify=True, allow_redirects=True, headers=headers, timeout=10)
            data = r.text
            data = data.replace(" ", "")
            buffer = io.StringIO(data)
            df = pd.read_csv(buffer, delimiter='\t')
            # Extract words to set for O(1) lookup
            if 'Orden' in df.columns:
                _FORMAS_5000_CACHE = set(df['Orden'].dropna().astype(str).str.lower().tolist())
            else:
                _FORMAS_5000_CACHE = set(df.iloc[:, 0].dropna().astype(str).str.lower().tolist())
        except Exception as e:
            print(f"Warning: Could not fetch 5000_formas.txt: {e}")
            _FORMAS_5000_CACHE = set()
    return _FORMAS_5000_CACHE

def interpretacion(orth_index, syll_index, synt_index, lex_index, red_index,extran_index):
    workbook = get_mapa_indices()
    worksheet = workbook.active
    progreso = 0
    aux_pauta1_frases_complejidad = -1
    aux_pauta1_frases_puntos = -1
    aux_pauta1_frases_ponderada = -1

    aux_pauta6_vocabulario_tris = -1
    aux_pauta6_vocabulario_polis = -1
    aux_pauta6_vocabulario_ext = -1

    #INDICES ORTOGRAFICOS

    #Puntos
    min_ptos_si = float(worksheet.cell(8, 4).value)
    max_ptos_si = float(worksheet.cell(8,8).value)
    min_ptos_no = float(worksheet.cell(8, 12).value)
    max_ptos_no = float(worksheet.cell(8,16).value)
    min_ptos_p1 = float(worksheet.cell(8, 20).value)
    max_ptos_p1 = float(worksheet.cell(8, 24).value)
    min_ptos_p2 = float(worksheet.cell(8, 28).value)
    max_ptos_p2 = float(worksheet.cell(8, 32).value)
    min_ptos_p3 = float(worksheet.cell(8, 36).value)
    max_ptos_p3 = float(worksheet.cell(8, 40).value)
    min_ptos_p4 = float(worksheet.cell(8, 44).value)
    max_ptos_p4 = float(worksheet.cell(8, 48).value)

    if orth_index['ind_puntos'] >= min_ptos_si and orth_index['ind_puntos'] <= max_ptos_si:
        aux_pauta1_frases_puntos = 1 #Se cumple el auxiliar
        pauta4_ortotipografia = 'green' #Se cumple la pauta
        progreso += 1
    elif orth_index['ind_puntos'] >= min_ptos_no and orth_index['ind_puntos'] <= max_ptos_no:
        pauta4_ortotipografia = 'red' #No se cumple la pauta
        aux_pauta1_frases_puntos = -1  # No se cumple el auxiliar
        progreso += 0
    elif orth_index['ind_puntos'] >= min_ptos_p1 and orth_index['ind_puntos'] <= max_ptos_p1:
        pauta4_ortotipografia = '#C70039' #Parcial1
        aux_pauta1_frases_puntos = -0.6  # Parcial 1
        progreso += 0.2
    elif orth_index['ind_puntos'] >= min_ptos_p2 and orth_index['ind_puntos'] <= max_ptos_p2:
        pauta4_ortotipografia = 'orange' #Parcial2
        aux_pauta1_frases_puntos = -0.4  # Parcial 2
        progreso += 0.4
    elif orth_index['ind_puntos'] >= min_ptos_p3 and orth_index['ind_puntos'] <= max_ptos_p3:
        pauta4_ortotipografia = '#FFC300' #Parcial3
        aux_pauta1_frases_puntos = 0.2  # Parcial 3
        progreso += 0.6
    elif orth_index['ind_puntos'] >= min_ptos_p4 and orth_index['ind_puntos'] <= max_ptos_p4:
        aux_pauta1_frases_puntos = 0.4  # Parcial 4
        pauta4_ortotipografia = '#57b400'  #Se cumple de forma parcial la pauta
        progreso += 0.8

    #Puntos y a parte
    min_ptos_parte_si = float(worksheet.cell(9, 4).value)
    max_ptos_parte_si = float(worksheet.cell(9, 8).value)
    min_ptos_parte_no = float(worksheet.cell(9, 12).value)
    max_ptos_parte_no = float(worksheet.cell(9, 16).value)
    min_ptos_parte_p1 = float(worksheet.cell(9, 20).value)
    max_ptos_parte_p1 = float(worksheet.cell(9, 24).value)
    min_ptos_parte_p2 = float(worksheet.cell(9, 28).value)
    max_ptos_parte_p2 = float(worksheet.cell(9, 32).value)
    min_ptos_parte_p3 = float(worksheet.cell(9, 36).value)
    max_ptos_parte_p3 = float(worksheet.cell(9, 40).value)
    min_ptos_parte_p4 = float(worksheet.cell(9, 44).value)
    max_ptos_parte_p4 = float(worksheet.cell(9, 48).value)
    if orth_index['ind_puntosaparte'] >= min_ptos_parte_si and orth_index['ind_puntosaparte'] <= max_ptos_parte_si:
        pauta3_ortotipografia = 'green' #Se cumple la pauta
        pauta5_ortotipografia = 'green'  # Se cumple la pauta
        progreso += 2
    elif orth_index['ind_puntosaparte'] >= min_ptos_parte_no and orth_index['ind_puntosaparte'] <= max_ptos_parte_no:
        pauta3_ortotipografia = 'red' #No se cumple la pauta
        pauta5_ortotipografia = 'red'  #Se cumple la pauta
    elif orth_index['ind_puntosaparte'] >= min_ptos_parte_p1 and orth_index['ind_puntosaparte'] <= max_ptos_parte_p1:
        pauta3_ortotipografia = '#C70039' #Parcial1
        pauta5_ortotipografia = '#C70039'  # Parcial1
        progreso += 0.4
    elif orth_index['ind_puntosaparte'] >= min_ptos_parte_p2 and orth_index['ind_puntosaparte'] <= max_ptos_parte_p2:
        pauta3_ortotipografia = 'orange' #Parcial2
        pauta5_ortotipografia = 'orange'  # Parcial2
        progreso += 0.8
    elif orth_index['ind_puntosaparte'] >= min_ptos_parte_p3 and orth_index['ind_puntosaparte'] <= max_ptos_parte_p3:
        pauta3_ortotipografia = '#FFC300'  # Parcial3
        pauta5_ortotipografia = '#FFC300'  # Parcial3
        progreso += 1.2
    elif orth_index['ind_puntosaparte'] >= min_ptos_parte_p4 and orth_index['ind_puntosaparte'] <= max_ptos_parte_p4:
        pauta3_ortotipografia = '#57b400'  # Parcial4
        pauta5_ortotipografia = '#57b400'  # Parcial4
        progreso += 1.6

    #Comas
    min_comas_si = float(worksheet.cell(10, 4).value)
    max_comas_si = float(worksheet.cell(10, 8).value)
    min_comas_no = float(worksheet.cell(10, 12).value)
    max_comas_no = float(worksheet.cell(10, 16).value)
    min_comas_p1 = float(worksheet.cell(10, 20).value)
    max_comas_p1 = float(worksheet.cell(10, 24).value)
    min_comas_p2 = float(worksheet.cell(10, 28).value)
    max_comas_p2 = float(worksheet.cell(10, 32).value)
    min_comas_p3 = float(worksheet.cell(10, 36).value)
    max_comas_p3 = float(worksheet.cell(10, 40).value)
    min_comas_p4 = float(worksheet.cell(10, 44).value)
    max_comas_p4 = float(worksheet.cell(10, 48).value)
    if orth_index['ind_comas'] >= min_comas_si and orth_index['ind_comas'] <= max_comas_si:
        pauta6_organizacion = 'green' #Se cumple la pauta
        pauta13_frases = 'green'  # Se cumple la pauta
        progreso += 2
    elif orth_index['ind_comas'] >= min_comas_no and orth_index['ind_comas'] <= max_comas_no:
        pauta6_organizacion = 'red' #No se cumple la pauta
        pauta13_frases = 'red'  #Se cumple la pauta
    elif orth_index['ind_comas'] >= min_comas_p1 and orth_index['ind_comas'] <= max_comas_p1:
        pauta6_organizacion = '#C70039' #Parcial1
        pauta13_frases = '#C70039'  # Parcial1
        progreso += 0.4
    elif orth_index['ind_comas'] >= min_comas_p2 and orth_index['ind_comas'] <= max_comas_p2:
        pauta6_organizacion = 'orange' #Parcial2
        pauta13_frases = 'orange'  # Parcial2
        progreso += 0.8
    elif orth_index['ind_comas'] >= min_comas_p3 and orth_index['ind_comas'] <= max_comas_p3:
        pauta6_organizacion = '#FFC300'  # Parcial3
        pauta13_frases = '#FFC300'  # Parcial3
        progreso += 1.2
    elif orth_index['ind_puntosaparte'] >= min_comas_p4 and orth_index['ind_puntosaparte'] <= max_comas_p4:
        pauta6_organizacion = '#57b400'  # Parcial4
        pauta13_frases = '#57b400'  # Parcial4
        progreso += 1.6
    else:
        pauta6_organizacion = 'red' #No se cumple la pauta
        pauta13_frases = 'red'  #Se cumple la pauta
        

    #Puntos y comas
    min_ptocomas_si = float(worksheet.cell(11, 4).value)
    max_ptocomas_si = float(worksheet.cell(11, 8).value)
    min_ptocomas_no = float(worksheet.cell(11, 12).value)
    max_ptocomas_no = float(worksheet.cell(11, 16).value)
    min_ptocomas_p1 = float(worksheet.cell(11, 20).value)
    max_ptocomas_p1 = float(worksheet.cell(11, 24).value)
    min_ptocomas_p2 = float(worksheet.cell(11, 28).value)
    max_ptocomas_p2 = float(worksheet.cell(11, 32).value)
    min_ptocomas_p3 = float(worksheet.cell(11, 36).value)
    max_ptocomas_p3 = float(worksheet.cell(11, 40).value)
    min_ptocomas_p4 = float(worksheet.cell(11, 44).value)
    max_ptocomas_p4 = float(worksheet.cell(11, 48).value)
    if orth_index['ind_puntoycoma'] >= min_ptocomas_si and orth_index['ind_puntoycoma'] <= max_ptocomas_si:
        pauta7_ortotipografia = 'green' #Se cumple la pauta
        progreso += 1
    elif orth_index['ind_puntoycoma'] >= min_ptocomas_no and orth_index['ind_puntoycoma'] <= max_ptocomas_no:
        pauta7_ortotipografia = 'red' #No se cumple la pauta
    elif orth_index['ind_puntoycoma'] >= min_ptocomas_p1 and orth_index['ind_puntoycoma'] <= max_ptocomas_p1:
        pauta7_ortotipografia = '#C70039'  # Parcial1
        progreso += 0.2
    elif orth_index['ind_puntoycoma'] >= min_ptocomas_p2 and orth_index['ind_puntoycoma'] <= max_ptocomas_p2:
        pauta7_ortotipografia = 'orange'  # Parcial2
        progreso += 0.4
    elif orth_index['ind_puntoycoma'] >= min_ptocomas_p3 and orth_index['ind_puntoycoma'] <= max_ptocomas_p3:
        pauta7_ortotipografia = '#FFC300'  # Parcial3
        progreso += 0.6
    elif orth_index['ind_puntoycoma'] >= min_ptocomas_p4 and orth_index['ind_puntoycoma'] <= max_ptocomas_p4:
        pauta7_ortotipografia = '#57b400'  # Parcial4
        progreso += 0.8


    #INDICES SILABICOS

    #Extension silabica
    min_extension_silabica_si = float(worksheet.cell(13, 4).value)
    max_extension_silabica_si = float(worksheet.cell(13, 8).value)
    min_extension_silabica_no = float(worksheet.cell(13, 12).value)
    max_extension_silabica_no = float(worksheet.cell(13, 16).value)
    min_extension_silabicas_p1 = float(worksheet.cell(13, 20).value)
    max_extension_silabica_p1 = float(worksheet.cell(13, 24).value)
    min_extension_silabica_p2 = float(worksheet.cell(13, 28).value)
    max_extension_silabica_p2 = float(worksheet.cell(13, 32).value)
    min_extension_silabica_p3 = float(worksheet.cell(13, 36).value)
    max_extension_silabica_p3 = float(worksheet.cell(13, 40).value)
    min_extension_silabica_p4 = float(worksheet.cell(13, 44).value)
    max_extension_silabica_p4 = float(worksheet.cell(13, 48).value)
    if syll_index['ind_extension_silabica'] >= min_extension_silabica_si and syll_index['ind_extension_silabica'] <= max_extension_silabica_si:
        aux_pauta6_vocabulario_ext = 1 #Se cumple la pauta auxiliar
    elif syll_index['ind_extension_silabica'] >= min_extension_silabica_no and syll_index['ind_extension_silabica'] <= max_extension_silabica_no:
        aux_pauta6_vocabulario_ext = -1 #No se cumple la pauta auxiliar
    elif syll_index['ind_extension_silabica'] >= min_extension_silabicas_p1 and syll_index['ind_extension_silabica'] <= max_extension_silabica_p1:
        aux_pauta6_vocabulario_ext = -0.6  # Parcial1
    elif syll_index['ind_extension_silabica'] >= min_extension_silabica_p2 and syll_index['ind_extension_silabica'] <= max_extension_silabica_p2:
        aux_pauta6_vocabulario_ext = -0.2  # Parcial2
    elif syll_index['ind_extension_silabica'] >= min_extension_silabica_p3 and syll_index['ind_extension_silabica'] <= max_extension_silabica_p3:
        aux_pauta6_vocabulario_ext = 0.4  # Parcial3
    elif syll_index['ind_extension_silabica'] >= min_extension_silabica_p4 and syll_index['ind_extension_silabica'] <= max_extension_silabica_p4:
        aux_pauta6_vocabulario_ext = 0.6 # Parcial4

    #Palabras trisilabas
    min_trisilabas_si = float(worksheet.cell(14, 4).value)
    max_trisilabas_si = float(worksheet.cell(14, 8).value)
    min_trisilabas_no = float(worksheet.cell(14, 12).value)
    max_trisilabas_no = float(worksheet.cell(14, 16).value)
    min_trisilabas_p1 = float(worksheet.cell(14, 20).value)
    max_trisilabas_p1 = float(worksheet.cell(14, 24).value)
    min_trisilabas_p2 = float(worksheet.cell(14, 28).value)
    max_trisilabas_p2 = float(worksheet.cell(14, 32).value)
    min_trisilabas_p3 = float(worksheet.cell(14, 36).value)
    max_trisilabas_p3 = float(worksheet.cell(14, 40).value)
    min_trisilabas_p4 = float(worksheet.cell(14, 44).value)
    max_trisilabas_p4 = float(worksheet.cell(14, 48).value)
    if syll_index['ind_palabras_tris'] >= min_trisilabas_si and syll_index['ind_palabras_tris'] <= max_trisilabas_si:
        aux_pauta6_vocabulario_tris = 1 #Se cumple la pauta auxiliar
    elif syll_index['ind_palabras_tris'] >= min_trisilabas_no and syll_index['ind_palabras_tris'] <= max_trisilabas_no:
        aux_pauta6_vocabulario_tris = -1 #No se cumple la pauta auxiliar
    elif syll_index['ind_palabras_tris'] >= min_trisilabas_p1 and syll_index['ind_palabras_tris'] <= max_trisilabas_p1:
        aux_pauta6_vocabulario_tris = -0.6  # Parcial1
    elif syll_index['ind_palabras_tris'] >= min_trisilabas_p2 and syll_index['ind_palabras_tris'] <= max_trisilabas_p2:
        aux_pauta6_vocabulario_tris = -0.2  # Parcial2
    elif syll_index['ind_palabras_tris'] >= min_trisilabas_p3 and syll_index['ind_palabras_tris'] <= max_trisilabas_p3:
        aux_pauta6_vocabulario_tris = 0.4  # Parcial3
    elif syll_index['ind_palabras_tris'] >= min_trisilabas_p4 and syll_index['ind_palabras_tris'] <= max_trisilabas_p4:
        aux_pauta6_vocabulario_tris = 0.6 # Parcial4

    #Palabras polisilabas
    min_polisilabas_si = float(worksheet.cell(15, 4).value)
    max_polisilabas_si = float(worksheet.cell(15, 8).value)
    min_polisilabas_no = float(worksheet.cell(15, 12).value)
    max_polisilabas_no = float(worksheet.cell(15, 16).value)
    min_polisilabas_p1 = float(worksheet.cell(15, 20).value)
    max_polisilabas_p1 = float(worksheet.cell(15, 24).value)
    min_polisilabas_p2 = float(worksheet.cell(15, 28).value)
    max_polisilabas_p2 = float(worksheet.cell(15, 32).value)
    min_polisilabas_p3 = float(worksheet.cell(15, 36).value)
    max_polisilabas_p3 = float(worksheet.cell(15, 40).value)
    min_polisilabas_p4 = float(worksheet.cell(15, 44).value)
    max_polisilabas_p4 = float(worksheet.cell(15, 48).value)
    if syll_index['ind_palabras_polis'] >= min_polisilabas_si and syll_index['ind_palabras_polis'] <= max_polisilabas_si:
        aux_pauta6_vocabulario_polis = 1 #Se cumple la pauta auxiliar
    elif syll_index['ind_palabras_polis'] >= min_polisilabas_no and syll_index['ind_palabras_polis'] <= max_polisilabas_no:
        aux_pauta6_vocabulario_polis = -1 #No se cumple la pauta auxiliar
    elif syll_index['ind_palabras_polis'] >= min_polisilabas_p1 and syll_index['ind_palabras_polis'] <= max_polisilabas_p1:
        aux_pauta6_vocabulario_polis = -0.6  # Parcial1
    elif syll_index['ind_palabras_polis'] >= min_polisilabas_p2 and syll_index['ind_palabras_polis'] <= max_polisilabas_p2:
        aux_pauta6_vocabulario_polis = -0.2  # Parcial2
    elif syll_index['ind_palabras_polis'] >= min_polisilabas_p3 and syll_index['ind_palabras_polis'] <= max_polisilabas_p3:
        aux_pauta6_vocabulario_polis = 0.4  # Parcial3
    elif syll_index['ind_palabras_polis'] >= min_polisilabas_p4 and syll_index['ind_palabras_polis'] <= max_polisilabas_p4:
        aux_pauta6_vocabulario_polis = 0.6  # Parcial4

    #Con los indices silabicos, componemos la valoracion final de la pauta 6 de vocabulario
    if aux_pauta6_vocabulario_ext == 1 and aux_pauta6_vocabulario_tris == 1 or aux_pauta6_vocabulario_polis == 1: #Si todas se cumplen totalmente
        pauta6_vocabulario = "green" #La pauta se cumple
        progreso += 1
    elif aux_pauta6_vocabulario_ext == -1 and aux_pauta6_vocabulario_tris == -1 and aux_pauta6_vocabulario_polis == -1: #Si todas se incumplen
        pauta6_vocabulario = "red" #La pauta no se cumple
    else:
        media = (aux_pauta6_vocabulario_ext + aux_pauta6_vocabulario_tris + aux_pauta6_vocabulario_polis)/3 #Calcula la media de los tres
        estimacion = valor_cercano(media) #Obtiene el valor mas proximo a la media
        if estimacion == 1:
            pauta6_vocabulario = "green"
            progreso += 1
        elif estimacion == -0.6: #Parcial 1
            pauta6_vocabulario = "#C70039"
            progreso += 0.2
        elif estimacion == -0.2: #Parcial 2
            pauta6_vocabulario = "orange"
            progreso += 0.4
        elif estimacion == 0.2: #Parcial 3
            pauta6_vocabulario = "#FFC300"
            progreso += 0.6
        elif estimacion == 0.6: #Parcial 4
            pauta6_vocabulario = "#57b400"
            progreso += 0.8
        else: #No cumple
            pauta6_vocabulario = "red"
            progreso += 0

    #INDICES LEXICOS

    #Diversidad
    min_diver_si = float(worksheet.cell(17, 4).value)
    max_diver_si = float(worksheet.cell(17, 8).value)
    min_diver_no = float(worksheet.cell(17, 12).value)
    max_diver_no = float(worksheet.cell(17, 16).value)
    min_diver_p1 = float(worksheet.cell(17, 20).value)
    max_diver_p1 = float(worksheet.cell(17, 24).value)
    min_diver_p2 = float(worksheet.cell(17, 28).value)
    max_diver_p2 = float(worksheet.cell(17, 32).value)
    min_diver_p3 = float(worksheet.cell(17, 36).value)
    max_diver_p3 = float(worksheet.cell(17, 40).value)
    min_diver_p4 = float(worksheet.cell(17, 44).value)
    max_diver_p4 = float(worksheet.cell(17, 48).value)
    if lex_index['ind_diversidad_palabras'] >= min_diver_si and lex_index['ind_diversidad_palabras'] <= max_diver_si:
        pauta17_vocabulario = 'green' #Se cumple la pauta
        progreso += 1
    elif lex_index['ind_diversidad_palabras'] >= min_diver_no and lex_index['ind_diversidad_palabras'] <= max_diver_no:
        pauta17_vocabulario = 'red' #No se cumple la pauta
    elif lex_index['ind_diversidad_palabras'] >= min_diver_p1 and lex_index['ind_diversidad_palabras'] <= max_diver_p1:
        pauta17_vocabulario = "#C70039" # Parcial1
        progreso += 0.2
    elif lex_index['ind_diversidad_palabras'] >= min_diver_p2 and lex_index['ind_diversidad_palabras'] <= max_diver_p2:
        pauta17_vocabulario = "orange" #Parcial2
        progreso += 0.4
    elif lex_index['ind_diversidad_palabras'] >= min_diver_p3 and lex_index['ind_diversidad_palabras'] <= max_diver_p3:
        pauta17_vocabulario = '#FFC300'  # Parcial3
        progreso += 0.6
    elif lex_index['ind_diversidad_palabras'] >= min_diver_p4 and lex_index['ind_diversidad_palabras'] <= max_diver_p4:
        pauta17_vocabulario = '#57b400'  # Parcial4
        progreso += 0.8

    #Frecuencia
    min_frec_si = float(worksheet.cell(18, 4).value)
    max_frec_si = float(worksheet.cell(18, 8).value)
    min_frec_no = float(worksheet.cell(18, 12).value)
    max_frec_no = float(worksheet.cell(18, 16).value)
    min_frec_p1 = float(worksheet.cell(18, 20).value)
    max_frec_p1 = float(worksheet.cell(18, 24).value)
    min_frec_p2 = float(worksheet.cell(18, 28).value)
    max_frec_p2 = float(worksheet.cell(18, 32).value)
    min_frec_p3 = float(worksheet.cell(18, 36).value)
    max_frec_p3 = float(worksheet.cell(18, 40).value)
    min_frec_p4 = float(worksheet.cell(18, 44).value)
    max_frec_p4 = float(worksheet.cell(18, 48).value)
    if lex_index['ind_frecuencia_lexica'] >= min_frec_si and lex_index['ind_frecuencia_lexica'] <= max_frec_si:
        pauta1_vocabulario = 'green' #Se cumple la pauta
        progreso += 1
    elif lex_index['ind_frecuencia_lexica'] >= min_frec_no and lex_index['ind_frecuencia_lexica'] <= max_frec_no:
        pauta1_vocabulario = 'red' #No se cumple la pauta
    elif lex_index['ind_frecuencia_lexica'] >= min_frec_p1 and lex_index['ind_frecuencia_lexica'] <= max_frec_p1:
        pauta1_vocabulario = "#C70039"  # Parcial1
        progreso += 0.2
    elif lex_index['ind_frecuencia_lexica'] >= min_frec_p2 and lex_index['ind_frecuencia_lexica'] <= max_frec_p2:
        pauta1_vocabulario = "orange"  # Parcial2
        progreso += 0.4
    elif lex_index['ind_frecuencia_lexica'] >= min_frec_p3 and lex_index['ind_frecuencia_lexica'] <= max_frec_p3:
        pauta1_vocabulario = '#FFC300'  # Parcial3
        progreso += 0.6
    elif lex_index['ind_frecuencia_lexica'] >= min_frec_p4 and lex_index['ind_frecuencia_lexica'] <= max_frec_p4:
        pauta1_vocabulario = '#57b400'  # Parcial4
        progreso += 0.8
    else: 
        pauta1_vocabulario= 'red'

    #INDICES SINTACTICOS

    #Complejidad oracional
    min_complejidad_si = float(worksheet.cell(20, 4).value)
    max_complejidad_si = float(worksheet.cell(20, 8).value)
    min_complejidad_no = float(worksheet.cell(20, 12).value)
    max_complejidad_no = float(worksheet.cell(20, 16).value)
    min_complejidad_p1 = float(worksheet.cell(20, 20).value)
    max_complejidad_p1 = float(worksheet.cell(20, 24).value)
    min_complejidad_p2 = float(worksheet.cell(20, 28).value)
    max_complejidad_p2 = float(worksheet.cell(20, 32).value)
    min_complejidad_p3 = float(worksheet.cell(20, 36).value)
    max_complejidad_p3 = float(worksheet.cell(20, 40).value)
    min_complejidad_p4 = float(worksheet.cell(20, 44).value)
    max_complejidad_p4 = float(worksheet.cell(20, 48).value)
    if synt_index['ind_global_complej_oracional'] >= min_complejidad_si and synt_index['ind_global_complej_oracional'] <= max_complejidad_si:
        aux_pauta1_frases_complejidad = 1 #Se cumple la pauta auxiliar
    elif synt_index['ind_global_complej_oracional'] >= min_complejidad_no and synt_index['ind_global_complej_oracional'] <= max_complejidad_no:
        aux_pauta1_frases_complejidad = -1 #No se cumple la pauta auxiliar
    elif synt_index['ind_global_complej_oracional'] >= min_complejidad_p1 and synt_index['ind_global_complej_oracional'] <= max_complejidad_p1:
        aux_pauta1_frases_complejidad = -0.6  # Parcial1
    elif synt_index['ind_global_complej_oracional'] >= min_complejidad_p2 and synt_index['ind_global_complej_oracional'] <= max_complejidad_p2:
        aux_pauta1_frases_complejidad = -0.2  # Parcial2
    elif synt_index['ind_global_complej_oracional'] >= min_complejidad_p3 and synt_index['ind_global_complej_oracional'] <= max_complejidad_p3:
        aux_pauta1_frases_complejidad = 0.4  # Parcial3
    elif synt_index['ind_global_complej_oracional'] >= min_complejidad_p4 and synt_index['ind_global_complej_oracional'] <= max_complejidad_p4:
        aux_pauta1_frases_complejidad = 0.6  # Parcial4

    #Complejidad ponderada
    min_ponderada_si = float(worksheet.cell(21, 4).value)
    max_ponderada_si = float(worksheet.cell(21, 8).value)
    min_ponderada_no = float(worksheet.cell(21, 12).value)
    max_ponderada_no = float(worksheet.cell(21, 16).value)
    min_ponderada_p1 = float(worksheet.cell(21, 20).value)
    max_ponderada_p1 = float(worksheet.cell(21, 24).value)
    min_ponderada_p2 = float(worksheet.cell(21, 28).value)
    max_ponderada_p2 = float(worksheet.cell(21, 32).value)
    min_ponderada_p3 = float(worksheet.cell(21, 36).value)
    max_ponderada_p3 = float(worksheet.cell(21, 40).value)
    min_ponderada_p4 = float(worksheet.cell(21, 44).value)
    max_ponderada_p4 = float(worksheet.cell(21, 48).value)
    if synt_index['ind_ponderado_complej_oracional'] >= min_ponderada_si and synt_index['ind_ponderado_complej_oracional'] <= max_ponderada_si:
        aux_pauta1_frases_ponderada = 1 #Se cumple la pauta auxiliar
    elif synt_index['ind_ponderado_complej_oracional'] >= min_ponderada_no and synt_index['ind_ponderado_complej_oracional'] <= max_ponderada_no:
        aux_pauta1_frases_ponderada = -1 #No se cumple la pauta auxiliar
    elif synt_index['ind_ponderado_complej_oracional'] >= min_ponderada_p1 and synt_index['ind_ponderado_complej_oracional'] <= max_ponderada_p1:
        aux_pauta1_frases_ponderada = -0.6  # Parcial1
    elif synt_index['ind_ponderado_complej_oracional'] >= min_ponderada_p2 and synt_index['ind_ponderado_complej_oracional'] <= max_ponderada_p2:
        aux_pauta1_frases_ponderada = -0.2  # Parcial2
    elif synt_index['ind_ponderado_complej_oracional'] >= min_ponderada_p3 and synt_index['ind_ponderado_complej_oracional'] <= max_ponderada_p3:
        aux_pauta1_frases_ponderada = 0.4  # Parcial3
    elif synt_index['ind_ponderado_complej_oracional'] >= min_ponderada_p4 and synt_index['ind_ponderado_complej_oracional'] <= max_ponderada_p4:
        aux_pauta1_frases_ponderada = 0.6  # Parcial4

    # Con los indices silabicos, componemos la valoracion final de la pauta 1 de frases

    if aux_pauta1_frases_puntos == 1 and aux_pauta1_frases_complejidad == 1 and aux_pauta1_frases_ponderada == 1: #Si todas se cumplen
        pauta1_frases = "green" #Se cumple la pauta
        progreso += 1
    elif aux_pauta1_frases_puntos == -1 and aux_pauta1_frases_complejidad == -1 and aux_pauta1_frases_ponderada == -1: #Si se incumplen todas
        pauta1_frases = "red" #No se cumple la pauta
    else:
        media = (aux_pauta1_frases_puntos + aux_pauta1_frases_complejidad + aux_pauta1_frases_ponderada) / 3  # Calcula la media de los tres
        estimacion = valor_cercano(media)  # Obtiene el valor mas proximo a la media
        if estimacion == 1:
            pauta1_frases = "green"
            progreso += 1
        elif estimacion == -0.6:  # Parcial 1
            pauta1_frases = "#C70039"
            progreso += 0.2
        elif estimacion == -0.2:  # Parcial 2
            pauta1_frases = "orange"
            progreso += 0.4
        elif estimacion == 0.2:  # Parcial 3
            pauta1_frases = "#FFC300"
            progreso += 0.6
        elif estimacion == 0.6:  # Parcial 4
            pauta1_frases = "#57b400"
            progreso += 0.8
        else:  # No cumple
            pauta1_frases = "red"
            progreso += 0


    def _color_label(color):
        mapping = {
            'green': 'Adecuado',
            '#57b400': 'Mayormente adecuado',
            '#FFC300': 'Parcialmente adecuado',
            'orange': 'Parcialmente inadecuado',
            '#C70039': 'Mayormente inadecuado',
            'red': 'Inadecuado'
        }
        return mapping.get(color, str(color))

    result = {
        'pauta3_ortotipografia' : _color_label(pauta3_ortotipografia),
        'pauta4_ortotipografia': _color_label(pauta4_ortotipografia),
        'pauta5_ortotipografia': _color_label(pauta5_ortotipografia),
        'pauta7_ortotipografia': _color_label(pauta7_ortotipografia),
        'pauta6_organizacion' : _color_label(pauta6_organizacion),
        'pauta1_frases': _color_label(pauta1_frases),
        'pauta13_frases' : _color_label(pauta13_frases),
        'pauta1_vocabulario' : _color_label(pauta1_vocabulario),
        'pauta6_vocabulario': _color_label(pauta6_vocabulario),
        'pauta17_vocabulario' : _color_label(pauta17_vocabulario),
        'progreso' : round(progreso*10,2),
    }
    return result

def valor_cercano(valor):
    lista = [-1, -0.6, -0.2, 0.2, 0.6, 1]
    array = numpy.asarray(lista)
    indice = (numpy.abs(array - valor)).argmin()
    return lista[indice]


def index_calculator(inputtxt, nlp):
    inputtxt = expand_txt(inputtxt)
    document = nlp(inputtxt)
    orth_index = orthographic_index(document)
    if orth_index is None:
        return None
    syll_index = syllabic_index(document)
    synt_index = syntactic_index(document, orth_index['word_count'])
    lex_index = lexic_index(document)
    red_index = readability_index(document)
    extran_index= extranjerismos_index(document)

    #Interpretar resultados

    interp = interpretacion(orth_index,syll_index,synt_index,lex_index,red_index,extran_index)

    result = {
        'indice_ortografico' : orth_index,
        'indice_silabico' : syll_index,
        'indice_lexico' : lex_index,
        'indice_sintactico' : synt_index,
        'indice_lecturabilidad' : red_index,
        'interpretacion' : interp,
        'indice_extranjerismos': extran_index
    }

    return result


def expand_txt(inputxt):
    if not inputxt: #Texto vacio
        return None

    resultado = inputxt
    abreviaturas = get_abreviaturas()

    for pattern, replacement in abreviaturas:
         resultado = pattern.sub(replacement, resultado)

    return resultado


def orthographic_index(document):
    wordcount = 0 #Contador de palabras
    stopcount = 0 #Contador de puntos
    newlinecount = 0 #Contador de puntos y a parte
    commacount = 0 #Contador de comas
    semicoloncount = 0 #Contador de puntos y comas
    ellipsiscount = 0 #Contador de puntos suspensivos

    dot = False  #Para identificar un punto como punto y a parte

    for token in document:
        if not token.is_punct and not token.is_space: #Si es una palabra
            wordcount += 1
        if token.text == '.':
            stopcount += 1
            dot = True
            continue
        if token.text == ',':
            commacount += 1
        if token.text == ';':
            semicoloncount += 1
        if  token.is_space and dot:
            newlinecount += 1
        if token.text == '...':
            ellipsiscount += 1
        dot = False

    if wordcount == 0: #Si el texto esta vacio
        return None

    result = {
        'ind_puntos' : round(stopcount/wordcount,3), #Indice de puntos
        'ind_puntosaparte' : round(newlinecount/wordcount,3), #Indice de puntos y aparte
        'ind_comas' : round(commacount/wordcount,3), #Indice de comas
        'ind_puntoycoma' : round(semicoloncount/wordcount,3), #Indice de puntos y comas
        'ind_puntossuspensivos' : round(ellipsiscount/wordcount, 3), #Indice de puntos suspensivos
        'word_count' : wordcount #Numero de palabras
    }

    return result


def syllabic_index(document):
    ntris = 0 #Num palabras trisilabas
    npolis = 0 #Num palabras polisilabas
    syl_count = 0 #Num de silabas de palabras lexicas
    lex_count = 0 #Num de palabras lexicas
    sil_frec=0
    #patron de silabas: CCVCC, CCV, CCVC,VCC,CVCC,
    SYLLABLE_PATTERNS = [r'(?![AEIOUaeiou])[A-Za-z](?![AEIOUaeiou])[A-Za-z][aeiouAEIOUáéíóúÁÉÍÓÚüÜ](?![AEIOUaeiou])[A-Za-z](?![AEIOUaeiou])[A-Za-z]',r'(?![AEIOUaeiou])[A-Za-z](?![AEIOUaeiou])[A-Za-z][aeiouAEIOUáéíóúÁÉÍÓÚüÜ]',r'[aeiouAEIOUáéíóúÁÉÍÓÚüÜ](?![AEIOUaeiou])[A-Za-z](?![AEIOUaeiou])[A-Za-z]',r'(?![AEIOUaeiou])[A-Za-z](?![AEIOUaeiou])[A-Za-z][aeiouAEIOUáéíóúÁÉÍÓÚüÜ](?![AEIOUaeiou])[A-Za-z]',r'[aeiouAEIOUáéíóúÁÉÍÓÚüÜ](?![AEIOUaeiou])[A-Za-z](?![AEIOUaeiou])[A-Za-z]',r'(?![AEIOUaeiou])[A-Za-z][aeiouAEIOUáéíóúÁÉÍÓÚüÜ](?![AEIOUaeiou])[A-Za-z](?![AEIOUaeiou])[A-Za-z]']
    


 
    

    for token in document:
        
        if not (token.is_punct and  token.is_space and token.text[-1] == '.'): #Si es una palabra y no es abreviatura
            try:

                aux = syllabify(token.text) #Aux contiene una lista de las silabas de la palabra actual
                lex_count += 1
                syl_count += len(aux)
                if len(aux) > 1:
                    npolis += 1
                if len(aux) == 3:
                    ntris += 1

                for syl in aux:
                    for pattern in SYLLABLE_PATTERNS:
                        if (re.search(pattern,syl)):
                            sil_frec +=1
                            break
              
            except Exception:
                aux = []
            
    if lex_count != 0:
        result = {
            'ind_extension_silabica' : round(syl_count/lex_count,3), #Extension silabica
            'ind_palabras_tris' : int(round(ntris/lex_count,3))*100,  #Indice de Palabras trisilabas
            'ind_palabras_polis' : int(round(npolis/lex_count,3)) *100, #Indice de Palabras polisilabas
            'ind_baja_frecuencia': round(sil_frec/syl_count,3),
        }
    else:
        result = {
            'ind_extension_silabica': 0,  # Extension silabica
            'ind_palabras_tris': 0,  # Indice de Palabras trisilabas
            'ind_palabras_polis': 0,  # Indice de Palabras polisilabas
            'ind_baja_frecuencia': 0,
        }

    return result


def syntactic_index(document, wordcount):
    sentencecount = 0
    propositioncount = 0
    simplecount = 0
    coordcount = 0
    coordcount2 = 0
    yuxtCount=0
    subcount = 0
    subordinada=0
    propsub=0
    propconj=0
    propyuxt=0

    for sentence in document.sents:
        sentencecount += 1
        cconj = 0
        prop = 0
        issub = False
        yuxt= False
        sconj=0
       

        for token in sentence:
            if token.tag_ == 'VERB' and token.dep_ != 'ROOT' and token.dep_ != 'AUX':
                propositioncount += 1
                prop += 1
            if token.tag_ == 'CCONJ':
                cconj += 1
                coordcount2+=1
                propconj+=1
            if token.tag_ == 'SCONJ':
                sconj += 1
                propsub+=1
            if token.tag_ == 'VERB' and (token.dep_ == 'csubj' or token.dep_ == 'ccomp' or token.dep_ == 'acl' or token.dep_ == 'advcl'):
                issub = True
                subordinada+=1
            if (str(token) == ',' or str(token) == ':' or str(token) == ';') :
                yuxtCount += 1
                yuxt= True
                propyuxt+=1
                

        if (yuxt==True and issub==True):
                
            yuxtCount-=1
            propyuxt+=1
            
            propyuxt-=1
            
            
           
            
        if cconj == prop: #Si todas las proposiciones estan unidas por nexos coordinantes se considera oracion coordinada
            coordcount += 1
        elif issub:
            subcount += 1
        else:
            simplecount += 1
       
    if subordinada==0:
        compOracional=0
       
    else:
            
        compOracional=(coordcount+ yuxtCount)/subordinada
        

    if propsub ==0:
        comProp=0
    else:
         comProp = (propconj + propyuxt)/propsub

       

       
    result = {
        'ind_palabras_frase' : round(wordcount/sentencecount,3),
        'ind_global_complej_oracional' : round(propositioncount/sentencecount,3),
        'ind_ponderado_complej_oracional' : round(((simplecount + 2*coordcount + 3*subcount)/sentencecount),3),
        'ind_complej_estruc_oracional': round (compOracional,3),
        'ind_complej_estruc_proposicional': round (comProp,3),
    }

    return result


def lexic_index(document):
    wordlist = [] #Lista de palabras unicas
    wordlistFrec=[] # lista de palabras poco frecuentes.
    wordcount = 0 #Contador de palabras
    lowfreccount = 0
    lex_count = 0
    
    formas_set = get_formas_5000_set()

    for token in document:
        if not token.is_punct and not token.is_space: #Si es una palabra en general
            wordcount += 1
            if not wordlist.__contains__(token.text.lower()): #Solo incluye palabras unicas
                wordlist.append(token.text.lower())

        is_frequent = token.text.lower() in formas_set

        if (not token.is_punct and not token.is_space) and not token.is_stop:  # Si es una palabra lexica
            lex_count += 1

        if (not token.is_punct and not token.is_space) and not is_frequent:  # Si es una palabra que no aparece en la lista
            lowfreccount += 1
            if not wordlistFrec.__contains__(token.text.lower()): #Solo incluye palabras unicas
                wordlistFrec.append(token.text.lower())
                

    if lex_count == 0:
        frec_lex = 0
    else:
        frec_lex = lowfreccount/lex_count

    if lowfreccount==0:
        var_index=0
    else:
        var_index= wordcount/len(wordlistFrec)

    result = {
        'ind_diversidad_palabras' : round(len(wordlist) / wordcount, 3), #Indice de diversidad de palabras
        'ind_frecuencia_lexica' : round(frec_lex,3), #Indice de frecuencia lexica
        'ind_variacion_vocab': round(var_index,3)
    }

    return result




def extranjerismos_index(document):
    wordcount = 0 #Contador de palabras
    foreingCount=0
    lex_count = 0
    encontrado=1
    
    palabras_set = get_palabras_todas()
    
    for token in document:
         if not token.is_punct and not token.is_space: #Si es una palabra en general
            wordcount += 1
            if str(token).lower() not in palabras_set:
                foreingCount+=1
        

       

    result = {
        'ind_cantidad_extranjerismos' : round(foreingCount / wordcount*100), #Indice de ratio de extranjerismos
    }

    return result



def _interpret_fernandez_huerta(L):
    if L < 30:
        return "muy difícil"
    elif L >= 30 and L < 50:
        return "difícil"
    elif L >= 50 and L < 60:
        return "algo difícil"
    elif L >= 60 and L < 70:
        return "normal"
    elif L >= 70 and L < 80:
        return "algo fácil"
    elif L >= 80 and L < 90:
        return "fácil"
    else:
        return "muy fácil"

def _interpret_szigriszt_pazos(P):
    if P <= 15:
        return "muy difícil"
    elif P > 15 and P <= 35:
        return "árido"
    elif P > 35 and P <= 50:
        return "bastante difícil"
    elif P > 50 and P <= 65:
        return "normal"
    elif P > 65 and P <= 75:
        return "bastante fácil"
    elif P > 75 and P <= 85:
        return "fácil"
    else:
        return "muy fácil"

def _interpret_inflesz(P):
    if P <= 40:
        return "muy difícil"
    elif P > 40 and P <= 55:
        return "algo difícil"
    elif P > 55 and P <= 65:
        return "normal"
    elif P > 65 and P <= 80:
        return "bastante fácil"
    else:
        return "muy fácil"

def _interpret_gutierrez(G):
    if G <= 33.33:
        return "difícil"
    if G > 33.33 and G < 66.66:
        return "normal"
    else:
        return "fácil"

def _interpret_mu(M):
    if M < 31:
        return "muy difícil"
    elif M >= 31 and M <= 51:
        return "difícil"
    elif M >= 51 and M < 61:
        return "un poco difícil"
    elif M >= 61 and M < 71:
        return "adecuado"
    elif M >= 71 and M < 81:
        return "un poco fácil"
    elif M >= 81 and M < 91:
        return "fácil"
    else:
        return "muy fácil"

def _calculate_mu_index(text):
    # Ported logic from legibilidad.mu
    words = ''.join(filter(lambda x: not x.isdigit(), text))
    clean = re.compile('\W+')
    words = clean.sub(' ', words).strip().split()
    n = len(words)
    if n <= 1:
        return 0
    word_lengths = [len(word) for word in words]
    try:
        mean = statistics.mean(word_lengths)
        variance = statistics.variance(word_lengths)
        if variance == 0:
            return 0
        mu_val = (n / (n - 1)) * (mean / variance) * 100
        return round(mu_val, 2)
    except Exception:
        return 0

def readability_index(document):
    text = document.text
    
    fh_index = textstat.fernandez_huerta(text)
    sz_index = textstat.szigriszt_pazos(text)
    mu_index = _calculate_mu_index(text)
    gu_index = textstat.gutierrez_polini(text)
    cr_index = textstat.crawford(text)

    result = {
        'fernandez_huerta_index' : fh_index,
        'fernandez_huerta_interpretacion' : _interpret_fernandez_huerta(fh_index),
        'inflesz_index' : sz_index,
        'inflesz_interpretacion' : _interpret_inflesz(sz_index),
        'szigriszt_pazos_interpretacion' : _interpret_szigriszt_pazos(sz_index),
        'mu_index' : mu_index,
        'mu_interpretacion' : _interpret_mu(mu_index),
        'polini_index' : gu_index,
        'polini_interpretacion' : _interpret_gutierrez(gu_index),
        'crawford_index' : cr_index,
    }
    return result




